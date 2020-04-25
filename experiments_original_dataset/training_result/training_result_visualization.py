import openpyxl
import numpy as np
from matplotlib import pyplot as plt
from scipy.interpolate import spline

model_list = ['vgg',
            'vgg_face_two_eyes',
            'vgg_rnn_face_two_eyes',
            'vgg_rnn_face_eye_mouth',
            'vgg_rnn_face_two_eyes_mouth']
loss_function_list = ['MAE', 'RMSE']
batch_size_list = [16, 20]
column_list = [5, 10, 11]

def plot_result_comparison(model_list, batch_size, loss_function, column):
    plt.style.use('seaborn')
    plt.figure()
    column_name = "metric"
    for model in model_list:
        print('processing model: ' + model)
        model_xlsx_name = model + '_training_result.xlsx'
        wb = openpyxl.load_workbook(model_xlsx_name)
        ws_name = "batch " + str(batch_size) + ' + ' + loss_function
        ws = wb[ws_name]
        r = 2
        column_epoch = 3

        epoch_list = []
        while ws.cell(r, column_epoch).value != None:
            epoch_list.append(ws.cell(r, column_epoch).value)
            r = r + 1

        column_name = ws.cell(1, column).value
        metric_list = []
        r = 2
        while ws.cell(r, column).value != None:
            metric_list.append(ws.cell(r, column).value)
            r = r + 1
        style = '-'
        if model == 'vgg':
            style = '-'
        elif model == 'vgg_face_two_eyes':
            style = '-'
        else:
            style = '--'

        epoch_list = np.array(epoch_list)
        metric_list = np.array(metric_list)

        epoch_list_new = np.linspace(epoch_list.min(), epoch_list.max(), 501)
        metric_list_new = spline(epoch_list, metric_list, epoch_list_new)
        plt.plot(epoch_list_new, metric_list_new, style)
    plt.legend(model_list, loc='lower right')
    plt.title('batch size = ' + str(batch_size) + ', loss function = ' + loss_function)
    plt.xlabel('epoch')
    plt.ylabel(column_name)
    plt.savefig('batch size = ' + str(batch_size) + ', loss function = ' + loss_function + ', ' + column_name + '.png')

if __name__ == '__main__':
    i = 1
    total = len(loss_function_list) * len(batch_size_list) * len(column_list)
    for column in column_list:
        for loss_function in loss_function_list:
            for batch_size in batch_size_list:
                    plot_result_comparison(model_list, batch_size, loss_function, column)
                    print("Finished: " + str(i) + '/' + str(total))
                    i = i + 1
