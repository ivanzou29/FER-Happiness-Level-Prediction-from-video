import openpyxl
import os
import numpy as numpy

batch_size_list = [16, 20]
loss_function_list = ['MAE', 'RMSE']
epoch_list = [50, 75, 100, 125, 150, 175, 200, 250, 300, 350, 400, 450, 500, 550, 600, 650, 700, 750, 800]
lr_list = [2e-05]

xlsx_name = 'vgg_rnn_face_eye_mouth_training_result.xlsx'
def find_2nd(string, substring):
   return string.find(substring, string.find(substring) + 1)

def fill_in(batch_size, loss_function, epoch, lr):
    print("Batch size = " + str(batch_size))
    print("Loss function = " + loss_function)
    print("Epoch = " + str(epoch))

    wb = openpyxl.load_workbook(xlsx_name)

    ws_name = 'batch ' + str(batch_size) + ' + ' + loss_function

    ws = wb[ws_name]
    main_dir_name = 'batch_size=' + str(batch_size) + ', ' + loss_function
    history_dir_name = 'training_history_BatchSize=' + str(batch_size) + '_Epoch=' + str(epoch) + '_FC_Num=20_TimeStep=8_DropOut=0.1_LearningRate=' + str(lr) + '_Loss=' + loss_function
    history_file = os.path.join(main_dir_name, history_dir_name, 'history.txt')

    f_history = open(history_file, 'r')
    lines = f_history.readlines()
    validation_loss_line = lines[-3]
    testing_loss_line = lines[-1]
    val_pos = validation_loss_line.find(':') + 2
    validation_loss = round(float(validation_loss_line[val_pos: val_pos+7]),4)
    test_pos = testing_loss_line.find(':') + 2
    testing_loss = round(float(testing_loss_line[test_pos: test_pos+7]),4)
    f_history.close()

    training_error_file = os.path.join(main_dir_name, history_dir_name, 'training_error_rate.txt')
    f_training_error = open(training_error_file, 'r')
    training_error_line = f_training_error.readline()
    training_error_1 = round(float(training_error_line[17:find_2nd(training_error_line, 'E')])/100,4)
    training_error_2 = round(float(training_error_line[find_2nd(training_error_line, ': ')+2:])/100,4)
    f_training_error.close()

    validation_error_file = os.path.join(main_dir_name, history_dir_name, 'validation_error_rate.txt')
    f_validation_error = open(validation_error_file, 'r')
    validation_error_line = f_validation_error.readline()
    validation_error_1 = round(float(validation_error_line[17:find_2nd(validation_error_line, 'E')])/100,4)
    validation_error_2 = round(float(validation_error_line[find_2nd(validation_error_line, ': ')+2:])/100,4)
    f_validation_error.close()

    testing_error_file = os.path.join(main_dir_name, history_dir_name, 'testing_error_rate.txt')
    f_testing_error = open(testing_error_file, 'r')
    testing_error_line = f_testing_error.readline()
    testing_error_1 = round(float(testing_error_line[17:find_2nd(testing_error_line, 'E')])/100,4)
    testing_error_2 = round(float(testing_error_line[find_2nd(testing_error_line, ': ')+2:])/100,4)
    f_testing_error.close()

    data_to_fill = [validation_loss, testing_loss, training_error_1, training_error_2, validation_error_1, validation_error_2, testing_error_1, testing_error_2]
    
    column_epoch = 3
    row = 2
    while (ws.cell(row, column_epoch).value != epoch):
        row += 1
    for i in range(len(data_to_fill)):
        ws.cell(row, column_epoch + 1 + i).value = data_to_fill[i]
    wb.save(xlsx_name)


if __name__ == '__main__':
    for batch_size in batch_size_list:
        for loss_function in loss_function_list:
            for epoch in epoch_list:
                for lr in lr_list:
                    fill_in(batch_size, loss_function, epoch, lr)
